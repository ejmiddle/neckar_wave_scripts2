from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


SOURCE_DIR = Path("workspace/Abzug Stunden")
OUTPUT_FILE = SOURCE_DIR / "Stunden_uebersicht.xlsx"


@dataclass(frozen=True)
class NormalizedRow:
    source_file: str
    date_text: str
    date: pd.Timestamp
    employee: str
    shift: str
    location: str
    worked_hours: float
    comment: str


def normalize_worked_hours(value: object) -> float:
    text = "" if value is None else str(value).strip()
    if not text:
        return float("nan")
    text = text.replace(",", ".")
    if re.fullmatch(r"\d+", text):
        raw_int = int(text)
        if raw_int >= 20 and raw_int % 10 == 5:
            return raw_int / 10.0
        return float(raw_int)
    return float(text)


def infer_worked_hours_from_date_text(date_text: object) -> float:
    text = "" if date_text is None else str(date_text).strip().strip('"')
    times = re.findall(r"\b(\d{1,2}):(\d{2})\b", text)
    if len(times) < 2:
        return float("nan")

    start_hours, start_minutes = (int(part) for part in times[0])
    end_hours, end_minutes = (int(part) for part in times[-1])
    start = start_hours + start_minutes / 60
    end = end_hours + end_minutes / 60
    if end < start:
        end += 24
    return round(end - start, 2)


def parse_date(date_text: object) -> pd.Timestamp:
    text = "" if date_text is None else str(date_text).strip().strip('"')
    match = re.match(r"^([A-Za-z]+ \d{1,2}, \d{4})", text)
    if not match:
        raise ValueError(f"Cannot parse date from: {text!r}")
    return pd.to_datetime(match.group(1), format="%B %d, %Y")


def load_csv_file(path: Path) -> pd.DataFrame:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = []
        for row in reader:
            normalized_row = {
                (key.strip() if isinstance(key, str) else key): value
                for key, value in row.items()
            }
            date_text = str(row.get("Date", "")).strip()
            if not date_text:
                continue
            employee = str(
                normalized_row.get("Mitarbeiter")
                or normalized_row.get("Property")
                or normalized_row.get("Employee")
                or ""
            ).strip()
            shift = str(normalized_row.get("Shift", "")).strip()
            location = str(normalized_row.get("Location", "")).strip()
            worked_hours = normalize_worked_hours(normalized_row.get("Worked Hours", ""))
            if pd.isna(worked_hours):
                worked_hours = infer_worked_hours_from_date_text(date_text)
            comment = str(normalized_row.get("Comment", "")).strip()
            if not employee and not shift and not location and pd.isna(worked_hours) and not comment:
                continue
            normalized = {
                "source_file": path.name,
                "date_text": date_text,
                "date": parse_date(date_text),
                "employee": employee,
                "shift": shift,
                "location": location,
                "worked_hours": worked_hours,
                "comment": comment,
            }
            rows.append(normalized)
    return pd.DataFrame(rows)


DEDUP_COLUMNS = [
    "date_text",
    "date",
    "employee",
    "shift",
    "location",
    "worked_hours",
    "comment",
]


def build_workbook() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    csv_files = sorted(SOURCE_DIR.glob("Zeiterfassung *.csv"))
    if not csv_files:
        raise FileNotFoundError(f"No CSV files found in {SOURCE_DIR}")

    frames = []
    audit_rows = []
    for path in csv_files:
        frame = load_csv_file(path)
        frames.append(frame)
        audit_rows.append(
            {
                "Datei": path.name,
                "Verarbeitete Eintraege": len(frame),
                "Von _all-Datei abgedeckt": "ja" if path.stem.endswith("_all") else "",
            }
        )

    loaded = pd.concat(frames, ignore_index=True)
    non_hour_rows = (
        loaded[loaded["worked_hours"].isna()]
        .drop_duplicates(subset=DEDUP_COLUMNS)
        .sort_values(["date", "employee", "shift", "location"])
        .reset_index(drop=True)
    )
    raw = loaded[loaded["worked_hours"].notna()].reset_index(drop=True)
    loaded_hours_count = len(raw)

    duplicate_rows = (
        raw.groupby(DEDUP_COLUMNS, dropna=False)
        .agg(
            Anzahl=("source_file", "size"),
            Quelldateien=("source_file", lambda values: ", ".join(sorted(set(values)))),
        )
        .reset_index()
    )
    duplicate_rows = duplicate_rows[duplicate_rows["Anzahl"] > 1].sort_values(
        ["date", "employee", "shift", "location"]
    )

    raw = (
        raw.drop_duplicates(subset=DEDUP_COLUMNS)
        .sort_values(["date", "employee", "shift", "location"])
        .reset_index(drop=True)
    )

    daily = (
        raw.pivot_table(
            index="date",
            columns="employee",
            values="worked_hours",
            aggfunc="sum",
            fill_value=0.0,
        )
        .sort_index()
    )
    daily.index.name = "Datum"
    daily["Gesamt"] = daily.sum(axis=1)
    daily = daily.reset_index()
    daily.insert(1, "Monat", daily["Datum"].dt.strftime("%Y-%m"))

    monthly = (
        raw.assign(Monat=raw["date"].dt.strftime("%Y-%m"))
        .pivot_table(
            index="Monat",
            columns="employee",
            values="worked_hours",
            aggfunc="sum",
            fill_value=0.0,
        )
        .sort_index()
        .reset_index()
    )
    monthly["Gesamt"] = monthly.drop(columns=["Monat"]).sum(axis=1)

    audit = pd.DataFrame(
        [
            {"Kennzahl": "CSV-Dateien", "Wert": len(csv_files)},
            {"Kennzahl": "Geladene Eintraege gesamt", "Wert": len(loaded)},
            {"Kennzahl": "Ausgeschlossene Nicht-Stunden-Eintraege eindeutig", "Wert": len(non_hour_rows)},
            {"Kennzahl": "Geladene Stunden-Eintraege vor Deduplizierung", "Wert": loaded_hours_count},
            {"Kennzahl": "Eindeutige Stunden-Eintraege", "Wert": len(raw)},
            {"Kennzahl": "Entfernte Stunden-Dubletten", "Wert": loaded_hours_count - len(raw)},
            {"Kennzahl": "Tage in Tagessummen", "Wert": len(daily)},
            {"Kennzahl": "Monate in Monatssummen", "Wert": len(monthly)},
            {"Kennzahl": "Gesamtstunden", "Wert": daily["Gesamt"].sum()},
        ]
    )
    source_audit = pd.DataFrame(audit_rows)
    audit = pd.concat([audit, pd.DataFrame([{}]), source_audit], ignore_index=True)

    return raw, daily, monthly, audit, duplicate_rows, non_hour_rows


def format_workbook(path: Path, sheets: list[str]) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    header_font = Font(bold=True)

    for sheet_name in sheets:
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for col_idx, column_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
            values = [cell.value for cell in column_cells[: min(len(column_cells), 200)]]
            max_len = max((len(str(v)) for v in values if v is not None), default=0)
            col_letter = get_column_letter(col_idx)
            header = ws[f"{col_letter}1"].value

            if header in {"Datum", "date"}:
                width = 13
                for cell in column_cells[1:]:
                    cell.number_format = "yyyy-mm-dd"
            elif header in {"worked_hours", "Gesamt"} or (
                values and all(isinstance(v, (int, float)) or v is None for v in values[1:])
            ):
                width = max(10, min(max_len + 2, 14))
                for cell in column_cells[1:]:
                    cell.number_format = "0.00"
            else:
                width = min(max_len + 2, 24)

            ws.column_dimensions[col_letter].width = width

    wb.save(path)


def main() -> None:
    raw, daily, monthly, audit, duplicate_rows, non_hour_rows = build_workbook()

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
        raw_out = raw.copy()
        raw_out["date"] = raw_out["date"].dt.date
        raw_out.to_excel(writer, sheet_name="Rohdaten", index=False)

        daily.to_excel(writer, sheet_name="Tagessummen", index=False)
        monthly.to_excel(writer, sheet_name="Monatssummen", index=False)
        audit.to_excel(writer, sheet_name="Pruefung", index=False)
        duplicate_rows.to_excel(writer, sheet_name="Dubletten", index=False)
        non_hour_rows.to_excel(writer, sheet_name="Nicht_Stunden", index=False)

    format_workbook(
        OUTPUT_FILE,
        ["Rohdaten", "Tagessummen", "Monatssummen", "Pruefung", "Dubletten", "Nicht_Stunden"],
    )

    print(f"Created {OUTPUT_FILE} with {len(raw)} unique rows")


if __name__ == "__main__":
    main()
