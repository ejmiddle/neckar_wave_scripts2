from __future__ import annotations

import json
from copy import deepcopy
from datetime import date, timedelta
from io import BytesIO
from pathlib import Path
import re
from typing import BinaryIO

from openpyxl import load_workbook

from src.sevdesk.constants import (
    MONTHLY_UMSATZ_TEMPLATE_PATH,
    VOUCHER_REDEEMED_TEMPLATE_PATH,
    VOUCHER_SOLD_TEMPLATE_PATH,
)
from src.sevdesk.voucher import normalize_create_payload

DEFAULT_SHEETS = ("ALT", "WIE")
EXPECTED_HEADER = (
    "Umsatz nach Steuersatz",
    "Umsatz final",
    "Umsatz inkl Steuer",
    "Rückerstattet inkl Steuer",
)
EXPECTED_SHEET_MARKERS = {
    "summe_umsatz_label": ("SUMME Umsatz", 4, 1),
    "voucher_verkauft_label": ("Voucher verkauft", 8, 1),
    "voucher_summe_label": ("Voucher Summe", 9, 1),
    "voucher_eingeloest_label": ("VOUCHER EINLÖS", 17, 1),
    "trinkgeld_label": ("Trinkgeld", 11, 1),
    "summe_umsatz_trinkgeld_label": ("SUMME Umsatz + Trinkgeld", 12, 1),
    "umsatz_7_label": ("Umsatz 7%", 2, 5),
    "umsatz_19_label": ("Umsatz 19%", 3, 5),
    "voucher_hinweis": ("Voucher werden nur Quartalsweise gebucht", 10, 5),
}
TARGET_RATES = ("0", "7", "19")
VOUCHER_RATES = ("7", "19")
REVENUE_TAX_RULE = {"id": 1, "objectName": "TaxRule"}
TEMPLATE_PATH = MONTHLY_UMSATZ_TEMPLATE_PATH
DEFAULT_REVENUE_ACCOUNTING_TYPE_NAME = "Einnahmen / Erlöse / Verkäufe"
DEFAULT_VOUCHER_ACCOUNTING_TYPE_NAME = "Verrechnungskonto Gutscheine"
MONTHLY_VOUCHER_TEMPLATE_MAP = {
    "voucher_verkauft": {
        "template_path": VOUCHER_SOLD_TEMPLATE_PATH,
        "description_prefix": "Voucher verkauft",
        "credit_debit": "D",
        "amount_key": "voucher_verkauft",
    },
    "voucher_eingeloest": {
        "template_path": VOUCHER_REDEEMED_TEMPLATE_PATH,
        "description_prefix": "Voucher eingelöst",
        "credit_debit": "C",
        "amount_key": "voucher_eingeloest",
    },
}
PLACEHOLDER_RE = re.compile(r"^\{\{\s*([a-zA-Z0-9_]+)\s*\}\}$")


class MonthlyUmsatzFormatError(ValueError):
    pass


def _normalize_rate(value: object) -> str | None:
    if value is None:
        return None

    if isinstance(value, str):
        cleaned = value.strip().replace("%", "").replace(",", ".")
        if not cleaned:
            return None
        try:
            number = float(cleaned)
        except ValueError:
            return None
    elif isinstance(value, (int, float)):
        number = float(value)
    else:
        return None

    if 0 <= number <= 1:
        number *= 100

    rounded = round(number)
    if abs(number - rounded) > 1e-9:
        return None

    return str(int(rounded))


def _find_header_row(sheet_name: str, worksheet: object) -> int:
    header_matches: list[int] = []

    for row_idx in range(1, worksheet.max_row + 1):
        row_values = tuple(worksheet.cell(row=row_idx, column=col).value for col in range(1, 5))
        if row_values == EXPECTED_HEADER:
            header_matches.append(row_idx)

    if not header_matches:
        raise MonthlyUmsatzFormatError(
            f"Sheet '{sheet_name}' is not in the expected format: Umsatz header row not found"
        )

    if len(header_matches) > 1:
        raise MonthlyUmsatzFormatError(
            f"Sheet '{sheet_name}' is not in the expected format: multiple Umsatz header rows found"
        )

    return header_matches[0]


def _sheet_signature(worksheet: object, header_row: int) -> dict[str, object]:
    signature = {
        "header_row": header_row,
        "max_column": worksheet.max_column,
    }

    for key, (_, row_offset, column) in EXPECTED_SHEET_MARKERS.items():
        signature[key] = worksheet.cell(row=header_row + row_offset, column=column).value

    return signature


def _validate_sheet_format(sheet_name: str, worksheet: object) -> tuple[int, dict[str, object]]:
    header_row = _find_header_row(sheet_name, worksheet)
    signature = _sheet_signature(worksheet, header_row)

    for _, (expected_value, row_offset, column) in EXPECTED_SHEET_MARKERS.items():
        actual_value = worksheet.cell(row=header_row + row_offset, column=column).value
        if actual_value != expected_value:
            raise MonthlyUmsatzFormatError(
                f"Sheet '{sheet_name}' is not in the expected format: expected "
                f"{expected_value!r} at row {header_row + row_offset}, column {column}, "
                f"got {actual_value!r}"
            )

    return header_row, signature


def _extract_sheet_values(sheet_name: str, worksheet: object) -> tuple[dict[str, float], dict[str, object]]:
    header_row, signature = _validate_sheet_format(sheet_name, worksheet)

    values: dict[str, float] = {}
    row_idx = header_row + 1

    for expected_rate in TARGET_RATES:
        rate_cell_value = worksheet.cell(row=row_idx, column=1).value
        rate = _normalize_rate(rate_cell_value)
        if rate != expected_rate:
            raise MonthlyUmsatzFormatError(
                f"Sheet '{sheet_name}' is not in the expected format: expected {expected_rate}% "
                f"in row {row_idx}, got {rate_cell_value!r}"
            )

        amount = worksheet.cell(row=row_idx, column=2).value
        if not isinstance(amount, (int, float)):
            raise MonthlyUmsatzFormatError(
                f"Sheet '{sheet_name}' is not in the expected format: expected numeric Umsatz "
                f"value for {expected_rate}% in row {row_idx}, got {amount!r}"
            )

        values[f"umsatz_{expected_rate}_prozent"] = float(amount)
        row_idx += 1

    voucher_sold_row = header_row + EXPECTED_SHEET_MARKERS["voucher_verkauft_label"][1]
    voucher_sold_amount = worksheet.cell(row=voucher_sold_row, column=2).value
    if not isinstance(voucher_sold_amount, (int, float)):
        raise MonthlyUmsatzFormatError(
            f"Sheet '{sheet_name}' is not in the expected format: expected numeric Voucher "
            f"verkauft value in row {voucher_sold_row}, got {voucher_sold_amount!r}"
        )
    values["voucher_verkauft"] = float(voucher_sold_amount)

    voucher_eingeloest_row = header_row + EXPECTED_SHEET_MARKERS["voucher_eingeloest_label"][1]
    voucher_eingeloest_amount = worksheet.cell(row=voucher_eingeloest_row, column=2).value
    if not isinstance(voucher_eingeloest_amount, (int, float)):
        raise MonthlyUmsatzFormatError(
            f"Sheet '{sheet_name}' is not in the expected format: expected numeric Voucher "
            f"eingelöst value in row {voucher_eingeloest_row}, got {voucher_eingeloest_amount!r}"
        )
    values["voucher_eingeloest"] = float(voucher_eingeloest_amount)

    return values, signature


def _open_workbook(workbook_source: str | Path | bytes | BinaryIO):
    if isinstance(workbook_source, (str, Path)):
        return load_workbook(workbook_source, data_only=True, read_only=True)
    if isinstance(workbook_source, bytes):
        return load_workbook(BytesIO(workbook_source), data_only=True, read_only=True)
    return load_workbook(workbook_source, data_only=True, read_only=True)


def extract_monthly_umsatz_json(
    workbook_source: str | Path | bytes | BinaryIO,
    sheet_names: tuple[str, ...] | list[str] | None = None,
) -> dict[str, dict[str, float]]:
    if sheet_names is None:
        sheet_names = list(DEFAULT_SHEETS)

    workbook = _open_workbook(workbook_source)

    missing_sheets = [sheet_name for sheet_name in sheet_names if sheet_name not in workbook.sheetnames]
    if missing_sheets:
        missing = ", ".join(missing_sheets)
        raise MonthlyUmsatzFormatError(f"Missing sheet(s) in workbook: {missing}")

    result: dict[str, dict[str, float]] = {}
    reference_signature: dict[str, object] | None = None
    reference_sheet: str | None = "ALT" if "ALT" in sheet_names else None

    if reference_sheet is not None:
        reference_values, reference_signature = _extract_sheet_values(
            reference_sheet,
            workbook[reference_sheet],
        )
        result[reference_sheet] = reference_values

    for sheet_name in sheet_names:
        if sheet_name == reference_sheet:
            continue

        values, signature = _extract_sheet_values(sheet_name, workbook[sheet_name])
        if reference_signature is None:
            reference_signature = signature
            reference_sheet = sheet_name
        elif signature != reference_signature:
            raise MonthlyUmsatzFormatError(
                f"Sheet '{sheet_name}' is not in the same format as '{reference_sheet}'"
            )
        result[sheet_name] = values

    return result


def previous_month_end(reference_date: date | None = None) -> date:
    target_date = reference_date or date.today()
    first_of_current_month = target_date.replace(day=1)
    return first_of_current_month - timedelta(days=1)


def format_german_date(value: date) -> str:
    return value.strftime("%d.%m.%Y")


def _round_amount(value: float) -> float:
    return round(float(value), 2)


def _compute_tax_components(sum_gross: float, tax_rate_percent: float) -> tuple[float, float, float]:
    gross = _round_amount(sum_gross)
    if tax_rate_percent == 0:
        return gross, gross, 0.0

    divisor = 1 + (tax_rate_percent / 100.0)
    net = _round_amount(gross / divisor)
    tax = _round_amount(gross - net)
    return gross, net, tax


def _load_voucher_template() -> dict[str, object]:
    if not TEMPLATE_PATH.exists():
        raise MonthlyUmsatzFormatError(f"Voucher template not found: {TEMPLATE_PATH}")

    try:
        payload = json.loads(TEMPLATE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise MonthlyUmsatzFormatError(f"Voucher template JSON is invalid: {exc}") from exc

    if not isinstance(payload, dict):
        raise MonthlyUmsatzFormatError("Voucher template must be a JSON object")

    return payload


def _load_template(template_path: Path) -> dict[str, object]:
    if not template_path.exists():
        raise MonthlyUmsatzFormatError(f"Voucher template not found: {template_path}")

    try:
        payload = json.loads(template_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise MonthlyUmsatzFormatError(f"Voucher template JSON is invalid: {exc}") from exc

    if not isinstance(payload, dict):
        raise MonthlyUmsatzFormatError("Voucher template must be a JSON object")
    return payload


def _render_template_value(value: object, context: dict[str, object]) -> object:
    if isinstance(value, str):
        match = PLACEHOLDER_RE.match(value.strip())
        if match:
            return context.get(match.group(1))
        return value
    if isinstance(value, list):
        return [_render_template_value(item, context) for item in value]
    if isinstance(value, dict):
        return {key: _render_template_value(item, context) for key, item in value.items()}
    return value


def _render_template_payload(template_path: Path, context: dict[str, object]) -> dict[str, object]:
    template = _load_template(template_path)
    rendered = _render_template_value(template, context)
    if not isinstance(rendered, dict):
        raise MonthlyUmsatzFormatError(f"Rendered voucher template is not an object: {template_path}")
    return rendered


def _select_default_revenue_accounting_type(
    accounting_type_rows: list[dict[str, object]] | None,
) -> dict[str, str]:
    if accounting_type_rows:
        for row in accounting_type_rows:
            if str(row.get("name", "")).strip() == DEFAULT_REVENUE_ACCOUNTING_TYPE_NAME:
                return {"id": str(row.get("id", "")), "objectName": "AccountingType"}

    return {"id": "26", "objectName": "AccountingType"}


def _select_default_voucher_accounting_type(
    accounting_type_rows: list[dict[str, object]] | None,
) -> dict[str, str]:
    if accounting_type_rows:
        for row in accounting_type_rows:
            if str(row.get("name", "")).strip() == DEFAULT_VOUCHER_ACCOUNTING_TYPE_NAME:
                return {"id": str(row.get("id", "")), "objectName": "AccountingType"}

    return {"id": "1120447", "objectName": "AccountingType"}


def _clean_voucher_template_for_create(voucher: dict[str, object]) -> dict[str, object]:
    cleaned = deepcopy(voucher)
    for field_name in ("id", "create", "update", "sevClient", "createUser"):
        cleaned.pop(field_name, None)
    return cleaned


def _build_monthly_voucher_payload(
    *,
    sheet_name: str,
    values: dict[str, float],
    belegdatum: date,
    accounting_type_rows: list[dict[str, object]] | None,
    template_path: Path,
    description_prefix: str,
    credit_debit: str,
    amount_key: str,
) -> dict[str, object]:
    amount = values.get(amount_key)
    if not isinstance(amount, (int, float)):
        raise MonthlyUmsatzFormatError(
            f"Missing extracted voucher value for {sheet_name} {amount_key}"
        )

    voucher_date = format_german_date(belegdatum)
    description = f"{description_prefix} {belegdatum.strftime('%m-%y')} {sheet_name}"
    rounded_amount = _round_amount(float(amount))
    template_context = {
        "voucher_id": None,
        "created_at": None,
        "updated_at": None,
        "belegdatum": voucher_date,
        "payment_deadline": voucher_date,
        "voucher_name": description,
        "sum_net": rounded_amount,
        "sum_tax": 0.0,
        "sum_gross": rounded_amount,
    }
    template_voucher = _render_template_payload(template_path, template_context)
    voucher = _clean_voucher_template_for_create(template_voucher)
    voucher.update(
        {
            "objectName": "Voucher",
            "mapAll": True,
            "voucherType": "VOU",
            "creditDebit": credit_debit,
            "status": 50,
            "currency": "EUR",
            "voucherDate": voucher_date,
            "deliveryDate": voucher_date,
            "paymentDeadline": voucher_date,
            "description": description,
            "sumNet": rounded_amount,
            "sumTax": 0.0,
            "sumGross": rounded_amount,
            "sumNetAccounting": rounded_amount,
            "sumTaxAccounting": 0.0,
            "sumGrossAccounting": rounded_amount,
            "taxType": "default",
        }
    )
    voucher.pop("payDate", None)
    voucher.pop("paidAmount", None)

    accounting_type = _select_default_voucher_accounting_type(accounting_type_rows)
    voucher_position = {
        "objectName": "VoucherPos",
        "mapAll": True,
        "net": False,
        "taxRate": 0.0,
        "sumGross": rounded_amount,
        "sumNet": rounded_amount,
        "comment": description,
        "accountingType": deepcopy(accounting_type),
    }

    voucher_payload = {
        "voucher": voucher,
        "voucherPosSave": [voucher_position],
        "voucherPosDelete": None,
        "filename": None,
        "notes": {
            "source": "monthly_umsatz_excel",
            "sheet": sheet_name,
            "belegdatum": voucher_date,
            "description": description,
            "amount": rounded_amount,
            "template_path": str(template_path),
            "accounting_type": accounting_type,
        },
    }
    return normalize_create_payload(voucher_payload)


def build_monthly_umsatz_voucher_payloads(
    extracted_data: dict[str, dict[str, float]],
    belegdatum: date,
    accounting_type_rows: list[dict[str, object]] | None = None,
) -> dict[str, dict[str, dict[str, object]]]:
    template_voucher = _load_voucher_template()
    revenue_accounting_type = _select_default_revenue_accounting_type(accounting_type_rows)
    voucher_date = format_german_date(belegdatum)
    payloads: dict[str, dict[str, dict[str, object]]] = {}

    for sheet_name, values in extracted_data.items():
        sheet_payloads: dict[str, dict[str, object]] = {}
        voucher = deepcopy(template_voucher)
        if not isinstance(voucher, dict):
            raise MonthlyUmsatzFormatError("Voucher template must resolve to an object")

        positions: list[dict[str, object]] = []
        total_gross = 0.0
        total_net = 0.0
        total_tax = 0.0

        for tax_rate in VOUCHER_RATES:
            gross_value = values.get(f"umsatz_{tax_rate}_prozent")
            if not isinstance(gross_value, (int, float)):
                raise MonthlyUmsatzFormatError(
                    f"Missing extracted Umsatz value for {sheet_name} {tax_rate}%"
                )

            tax_rate_percent = float(tax_rate)
            gross, net, tax = _compute_tax_components(float(gross_value), tax_rate_percent)
            total_gross += gross
            total_net += net
            total_tax += tax
            positions.append(
                {
                    "objectName": "VoucherPos",
                    "mapAll": True,
                    "net": False,
                    "taxRate": tax_rate_percent,
                    "sumGross": gross,
                    "sumNet": net,
                    "comment": f"Umsatz {sheet_name} {tax_rate}%",
                    "accountingType": deepcopy(revenue_accounting_type),
                    "taxRule": dict(REVENUE_TAX_RULE),
                }
            )

        total_gross = _round_amount(total_gross)
        total_net = _round_amount(total_net)
        total_tax = _round_amount(total_tax)
        description = f"Umsatz {belegdatum.strftime('%m-%y')} {sheet_name}"

        voucher.update(
            {
                "objectName": "Voucher",
                "mapAll": True,
                "voucherType": "VOU",
                "creditDebit": "D",
                "status": 50,
                "currency": "EUR",
                "voucherDate": voucher_date,
                "deliveryDate": voucher_date,
                "paymentDeadline": voucher_date,
                "description": description,
                "sumNet": total_net,
                "sumTax": total_tax,
                "sumGross": total_gross,
                "sumNetAccounting": total_net,
                "sumTaxAccounting": total_tax,
                "sumGrossAccounting": total_gross,
                "taxRule": dict(REVENUE_TAX_RULE),
                "taxType": "default",
            }
        )
        voucher.pop("id", None)
        voucher.pop("create", None)
        voucher.pop("update", None)

        sheet_payloads["umsatz"] = {
            "voucher": voucher,
            "voucherPosSave": positions,
            "voucherPosDelete": None,
            "filename": None,
            "notes": {
                "source": "monthly_umsatz_excel",
                "sheet": sheet_name,
                "belegdatum": voucher_date,
                "extracted_umsatz": values,
                "template_path": str(TEMPLATE_PATH),
            },
        }
        for voucher_key, template_config in MONTHLY_VOUCHER_TEMPLATE_MAP.items():
            sheet_payloads[voucher_key] = _build_monthly_voucher_payload(
                sheet_name=sheet_name,
                values=values,
                belegdatum=belegdatum,
                accounting_type_rows=accounting_type_rows,
                template_path=Path(template_config["template_path"]),
                description_prefix=str(template_config["description_prefix"]),
                credit_debit=str(template_config["credit_debit"]),
                amount_key=str(template_config["amount_key"]),
            )

        payloads[sheet_name] = sheet_payloads

    return payloads
