from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

from src.accounting.finom_open_payments import build_finom_open_payments_result


def test_build_finom_open_payments_result_exports_compact_columns_and_owner() -> None:
    open_payments_csv = "\n".join(
        [
            '"Status";"Name";"Beschreibung";"Bezahldatum";"Betrag";',
            '"offen";"WAGNERMUEHLE.DE";"WAGNERMUEHLE.DE / MCC: 5411";"19.05.2026";"-22,70";',
            '"offen";"Neckar Wave Foods GmbH";"Transfer";"12.05.2026";"531,46";',
        ]
    ).encode()
    finom_statement_csv = "\n".join(
        [
            "Buchungsdatum,Time completed,Status,Transaktionsart,Auftraggeber/Empfänger,Counterparty BIC,Counterparty IBAN,Verwendungszweck,Tags,Zahlungsfreigeber,Kartennummer,Ursprungswährung,Ursprungsbetrag,Zahlungswährung,Zahlungsbetrag,Wallet-Saldo nach Transaktion,Wallet-Name,Wallet-IBAN,Begleitende Dokumente,Transaktions-ID",
            "19.05.2026,09:14,Completed,Card,WAGNERMUEHLE.DE,,,N/A,GROCERY,encrypted,***0937,EUR,-22.70,EUR,-22.70,711.89,Main,DE51100180000933228565,N/A,tx-card",
            "12.05.2026,15:24,Completed,Transfer,Neckar Wave Foods GmbH,,,Invoice,N/A,,,"
            "EUR,531.46,EUR,531.46,1000.00,Main,DE51100180000933228565,N/A,tx-transfer",
        ]
    ).encode()

    result = build_finom_open_payments_result(open_payments_csv, finom_statement_csv)

    assert result.enriched.columns[0] == "Finom Karteninhaber"
    assert result.enriched.loc[0, "Finom Karteninhaber"] == "Andi"
    assert result.enriched.loc[1, "Finom Karteninhaber"] == ""
    assert "Match Status" not in result.enriched.columns
    assert "Match Reason" not in result.enriched.columns
    assert "Finom Buchungsdatum" not in result.enriched.columns
    assert "Finom Time completed" not in result.enriched.columns
    assert "Finom Counterparty IBAN" not in result.enriched.columns
    assert "Finom Wallet-Name" not in result.enriched.columns
    assert result.xlsx_bytes.startswith(b"PK")
    assert list(result.owner_summary["Finom Karteninhaber"]) == ["Ohne Karte", "Andi"]

    workbook = load_workbook(BytesIO(result.xlsx_bytes), data_only=True)
    worksheet = workbook["Open payments enriched"]
    headers = [cell.value for cell in worksheet[1]]
    date_cell = worksheet.cell(row=2, column=headers.index("Open Bezahldatum") + 1)
    assert isinstance(date_cell.value, datetime)
    assert date_cell.number_format == "DD.MM.YYYY"
