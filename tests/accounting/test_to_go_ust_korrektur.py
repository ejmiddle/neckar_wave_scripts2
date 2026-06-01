from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook

from src.accounting.to_go_ust_korrektur import (
    ToGoUstKorrekturError,
    analyze_to_go_ust_korrektur_csv,
    create_to_go_ust_korrektur_workbook,
)


def test_to_go_ust_korrektur_groups_variants_and_writes_workbook() -> None:
    csv_text = "\n".join(
        [
            '"rechnung_nummer";"artikel_bezeichnung";"artikel_menge";"artikel_summe";"warengruppe_bezeichnung"',
            '"RG1";"Flat White (TO GO, Kuh)";"1,0000";"4,30";"Hot drinks"',
            '"RG2";"Flat White (TO GO, Kuh)";"2,0000";"8,60";"Hot drinks"',
            '"RG3";"Latte (Hafer, TO GO)";"1,0000";"4,50";"Hot drinks"',
            '"RG4";"Pizza stk (TO GO)";"1,0000";"3,23";"Snacks"',
            '"RG5";"Flat White (Kuh)";"1,0000";"4,30";"Hot drinks"',
            '"RG6";"Latte (Decaf, TO GO, Kuh)";"-1,0000";"-4,70";"Hot drinks"',
        ]
    )

    result = analyze_to_go_ust_korrektur_csv(csv_text.encode("latin1"))

    assert result.overview.to_dict("records") == [
        {
            "gruppe": "TO GO + Kuh",
            "varianten": 2,
            "zeilen": 3,
            "menge_summe": 2.0,
            "brutto_summe": 8.2,
        },
        {
            "gruppe": "TO GO ohne Kuh",
            "varianten": 2,
            "zeilen": 2,
            "menge_summe": 2.0,
            "brutto_summe": 7.73,
        },
    ]
    assert result.to_go_kuh_summary.to_dict("records") == [
        {
            "artikel_bezeichnung": "Flat White (TO GO, Kuh)",
            "zeilen": 2,
            "menge_summe": 3.0,
            "brutto_summe": 12.9,
        },
        {
            "artikel_bezeichnung": "Latte (Decaf, TO GO, Kuh)",
            "zeilen": 1,
            "menge_summe": -1.0,
            "brutto_summe": -4.7,
        },
    ]
    assert result.to_go_without_kuh_summary["artikel_bezeichnung"].tolist() == [
        "Latte (Hafer, TO GO)",
        "Pizza stk (TO GO)",
    ]

    workbook_bytes = create_to_go_ust_korrektur_workbook(result)
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True)

    assert workbook.sheetnames == [
        "Overview",
        "TO GO + Kuh summary",
        "TO GO no Kuh summary",
        "TO GO + Kuh rows",
        "TO GO no Kuh rows",
    ]
    assert workbook["Overview"]["A2"].value == "TO GO + Kuh"


def test_to_go_ust_korrektur_reports_missing_required_columns() -> None:
    csv_text = '"artikel_bezeichnung";"artikel_summe"\n"Flat White (TO GO, Kuh)";"4,30"'

    with pytest.raises(ToGoUstKorrekturError, match="artikel_menge"):
        analyze_to_go_ust_korrektur_csv(csv_text.encode("utf-8"))


def test_to_go_ust_korrektur_reports_non_numeric_values() -> None:
    csv_text = "\n".join(
        [
            '"artikel_bezeichnung";"artikel_menge";"artikel_summe"',
            '"Flat White (TO GO, Kuh)";"1,0000";"abc"',
        ]
    )

    with pytest.raises(ToGoUstKorrekturError, match="artikel_summe"):
        analyze_to_go_ust_korrektur_csv(csv_text.encode("utf-8"))
